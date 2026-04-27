# ── IMPORTS ────────────────────────────────────────────────────────────────
import os
from dotenv import load_dotenv
# Load environment variables from .env file
load_dotenv()

# Load Flask and web-related tools
from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, make_response, abort
# Load SQLAlchemy for database operations
from flask_sqlalchemy import SQLAlchemy
# Load LoginManager to handle user sessions (login/logout)
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
# wraps: helper for creating decorators (like role_required)
from functools import wraps
# Import our models: the blueprint for our database tables
from models import db, User, Vendor, Receipt, Street, FeeSchedule, FeePayment, Fine
# Load time handling libraries
from datetime import datetime, date, timedelta
# Load OS library for file paths
import os
# Load io for memory-based file handling (PDFs/Excel)
import io

# ── Optional Libraries (ReportLab for PDFs) ──────────────────────────────────
try:
    from reportlab.lib.pagesizes import A5
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False # Fallback if library is missing

# ── Optional Libraries (OpenPyXL for Excel) ─────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Optional Libraries (Africa's Talking for SMS) ──────────────────────────
try:
    import africastalking
    AT_USERNAME = os.environ.get('AT_USERNAME', 'sandbox')
    AT_API_KEY  = os.environ.get('AT_API_KEY',  'atsk_sandbox_key_placeholder')
    AT_SENDER   = os.environ.get('AT_SENDER',   '')
    AFRICASTALKING_OK = True
except ImportError:
    AFRICASTALKING_OK = False

# ── APP CONFIGURATION ───────────────────────────────────────────────────────
app = Flask(__name__) # Create the Flask application object
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'arua-vendor-dev-key') # Use ENV secret if available

# DATABASE CONFIGURATION (Dynamic for Local vs Production)
db_url = os.environ.get('DATABASE_URL')
if db_url and db_url.startswith("postgres://"):
    # Fix for newer SQLAlchemy versions that require 'postgresql://' instead of 'postgres://'
    db_url = db_url.replace("postgres://", "postgresql://", 1)

# Fallback to local SQLite if no external DATABASE_URL is provided
app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///' + os.path.join(app.instance_path, 'vendors_v2.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False # Performance optimization

db.init_app(app) # Connect database object to this app

# ── SERVERLESS INITIALIZATION ──────────────────────────────────────────────
# On Vercel, the 'if __name__ == "__main__"' block doesn't run.
# This hook ensures the database is created and seeded on the first request.
@app.before_request
def initialize_database():
    """Create tables and seed admin data if it's the first run."""
    # We use a global variable or check for table existence to avoid repeated runs
    if not hasattr(app, '_db_initialized'):
        with app.app_context():
            db.create_all()
            seed_demo_data() # Add admin/streets/fees
            app._db_initialized = True

login_manager = LoginManager(app) # Setup login manager
login_manager.login_view = 'login' # Where to redirect if user isn't logged in
login_manager.login_message = 'Please log in to access the system.' # Custom login message

@login_manager.user_loader
def load_user(user_id):
    """How Flask-Login finds a user by their ID."""
    return User.query.get(int(user_id))


# ── RBAC SECURITY DECORATOR ────────────────────────────────────────────────
def role_required(*roles):
    """Decorator to restrict access to specific roles (e.g. Admin only)."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # If user's role isn't in the allowed list, deny access
            if current_user.role not in roles and current_user.role != 'admin':
                abort(403) # Forbidden error
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ── AUTHENTICATION ROUTES ──────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login."""
    if current_user.is_authenticated: # Already logged in? Skip to dashboard
        return redirect(url_for('dashboard'))
    if request.method == 'POST': # If form submitted
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first() # Find user in DB
        if user and user.check_password(password): # Verify password hash
            login_user(user) # Start the session
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'danger') # Error feedback
    return render_template('login.html') # Show login page

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    """Allow new users to register."""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username')
        email    = request.form.get('email')
        password = request.form.get('password')
        role     = request.form.get('role', 'field_officer') # Default to field officer
        
        # Check if username or email already exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
        elif User.query.filter_by(email=email).first():
            flash('Email already registered.', 'danger')
        else:
            try:
                new_user = User(username=username, email=email, role=role)
                new_user.set_password(password) # Hash password
                new_user.created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                db.session.add(new_user) # Queue for database
                db.session.commit() # Save to database
                flash('Account created! Please log in.', 'success')
                return redirect(url_for('login'))
            except Exception as e:
                db.session.rollback() # Undo if error occurs
                flash(f'Error creating account: {str(e)}', 'danger')
    return render_template('signup.html')

@app.route('/logout')
@login_required # User must be logged in to log out
def logout():
    """Log the user out."""
    logout_user()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))


# ── DASHBOARD LOGIC ────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    """Main landing page with stats and role-specific views."""
    # Run background check for payment statuses
    update_vendor_payment_status()

    # If the user is a Vendor, show them their personal portal
    if current_user.role == 'vendor' and current_user.vendor_ptr:
        vendor = Vendor.query.get(current_user.vendor_ptr)
        my_receipts = Receipt.query.filter_by(vendor_id=vendor.id).order_by(Receipt.id.desc()).all()
        my_fines    = Fine.query.filter_by(vendor_id=vendor.id).all()
        
        total_paid  = sum(r.amount for r in my_receipts)
        unpaid_fines = [f for f in my_fines if f.status == 'Unpaid']
        
        return render_template('vendor_portal.html',
            vendor=vendor,
            receipts=my_receipts[:5], # Show last 5 receipts
            fines=my_fines[:5], # Show last 5 fines
            total_paid=total_paid,
            unpaid_fines_count=len(unpaid_fines),
            unpaid_fines_total=sum(f.amount for f in unpaid_fines)
        )

    # General Dashboard for Admin/Finance/Field Officers
    total_vendors   = Vendor.query.count()
    active_vendors  = Vendor.query.filter_by(status='Active').count()
    not_paid_count  = Vendor.query.filter_by(status='NotPaid').count()
    total_receipts  = Receipt.query.count()
    unverified_count = Receipt.query.filter_by(is_verified=False).count() # For Finance Officers
    
    today_str       = date.today().strftime('%Y-%m-%d')
    # Calculate today's revenue from receipts table
    today_revenue   = db.session.query(db.func.sum(Receipt.amount))\
                        .filter(Receipt.date_issued.like(today_str + '%')).scalar() or 0
    
    # Weekly revenue calculation (Monday to today)
    today = date.today()
    start_of_week = today - timedelta(days=today.weekday())
    weekly_revenue = db.session.query(db.func.sum(Receipt.amount))\
                        .filter(Receipt.date_issued >= start_of_week.strftime('%Y-%m-%d')).scalar() or 0

    total_revenue   = db.session.query(db.func.sum(Receipt.amount)).scalar() or 0
    recent_receipts = Receipt.query.order_by(Receipt.id.desc()).limit(5).all() # Latest 5 receipts
    recent_vendors  = Vendor.query.order_by(Vendor.id.desc()).limit(5).all() # Latest 5 registrations

    # Breakdown by trade type for the dashboard chart
    from sqlalchemy import func
    trade_data = db.session.query(Vendor.trade_type, func.count(Vendor.id))\
                            .group_by(Vendor.trade_type).all()

    # Find vendors whose permits expire in the next 30 days
    today_dt   = date.today()
    expiring_soon = []
    all_vendors = Vendor.query.filter(Vendor.status != 'Expired').all()
    for v in all_vendors:
        if v.permit_end:
            try:
                exp = date.fromisoformat(v.permit_end)
                days_left = (exp - today_dt).days
                if 0 <= days_left <= 30:
                    expiring_soon.append({'vendor': v, 'days_left': days_left, 'permit_end': v.permit_end})
            except ValueError:
                pass
    expiring_soon.sort(key=lambda x: x['days_left'])

    return render_template('dashboard.html',
        total_vendors=total_vendors,
        active_vendors=active_vendors,
        not_paid_count=not_paid_count,
        total_receipts=total_receipts,
        unverified_count=unverified_count,
        today_revenue=today_revenue,
        weekly_revenue=weekly_revenue,
        total_revenue=total_revenue,
        recent_receipts=recent_receipts,
        recent_vendors=recent_vendors,
        trade_data=trade_data,
        expiring_soon=expiring_soon,
        now=datetime.now()
    )


def update_vendor_payment_status():
    """Flags vendors as 'NotPaid' if they missed today's daily fee."""
    today_str = date.today().strftime('%Y-%m-%d')
    daily_schedules = FeeSchedule.query.filter_by(frequency='daily').all()
    
    if not daily_schedules:
        return

    active_vendors = Vendor.query.filter(Vendor.status.in_(['Active', 'NotPaid'])).all()
    
    for v in active_vendors:
        paid_today = False
        for sched in daily_schedules:
            # Look for a payment record matching today and this vendor
            payment = FeePayment.query.filter_by(
                vendor_id=v.id, 
                schedule_id=sched.id, 
                period_date=today_str,
                status='Paid'
            ).first()
            if payment:
                paid_today = True
                break
        
        if not paid_today:
            v.status = 'NotPaid' # Automatically set status to NotPaid
        else:
            v.status = 'Active' # Set back to Active if paid
    
    db.session.commit() # Save all status changes


# ── VENDOR REGISTRY ─────────────────────────────────────────────────────────
@app.route('/vendors')
@login_required
def vendors():
    """Displays a searchable list of all registered vendors."""
    search  = request.args.get('search', '')
    status  = request.args.get('status', '')
    query   = Vendor.query
    if search:
        # Search by name, ID, or street (joining Street table for better matching)
        query = query.join(Street, Vendor.street_id == Street.id, isouter=True).filter(
            db.or_(
                Vendor.full_name.ilike(f'%{search}%'),
                Vendor.vendor_id.ilike(f'%{search}%'),
                Vendor.street.ilike(f'%{search}%'),
                Street.name.ilike(f'%{search}%')
            )
        )
    if status:
        query = query.filter(Vendor.status == status) # Filter by status dropdown
    vendor_list = query.order_by(Vendor.id.desc()).all()
    return render_template('vendors.html', vendors=vendor_list, search=search, status=status)


@app.route('/vendors/<int:vendor_id>')
@login_required
def vendor_detail(vendor_id):
    """Shows the full profile and history of a specific vendor."""
    vendor   = Vendor.query.get_or_404(vendor_id)
    receipts = Receipt.query.filter_by(vendor_id=vendor.id).order_by(Receipt.id.desc()).all()
    return render_template('vendor_detail.html', vendor=vendor, receipts=receipts)


@app.route('/users', methods=['GET', 'POST'])
@login_required
@role_required('admin') # Admin only!
def manage_users():
    """Admin-only portal to manage system accounts."""
    if request.method == 'POST':
        username = request.form['username']
        email    = request.form['email']
        role     = request.form['role']
        password = request.form['password']
        vendor_id = request.form.get('vendor_id', type=int)
        
        # Check for uniqueness before creating
        existing = User.query.filter((User.username==username) | (User.email==email)).first()
        if existing:
            flash('Username or email already exists.', 'danger')
            return redirect(url_for('manage_users'))
            
        new_user = User(username=username, email=email, role=role)
        if role == 'vendor' and vendor_id:
            new_user.vendor_ptr = vendor_id # Link to vendor data if role is 'vendor'
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        flash(f'User {username} created successfully!', 'success')
        return redirect(url_for('manage_users'))
        
    users = User.query.all()
    vendors = Vendor.query.order_by(Vendor.full_name).all()
    return render_template('manage_users.html', users=users, vendors=vendors)



# ── VENDOR REGISTRATION ──────────────────────────────────────────────────────
@app.route('/vendors/register', methods=['GET', 'POST'])
@login_required # Login required to register a vendor
@role_required('admin', 'field_officer') # Only Admin or Field Officer can register
def register_vendor():
    """Handles the form to register a new vendor."""
    streets = Street.query.order_by(Street.name).all() # Get list of streets for dropdown
    if request.method == 'POST': # If form is submitted
        full_name   = request.form.get('full_name', '').strip() # Get name
        phone       = request.form.get('phone', '').strip() # Get phone
        street_name = request.form.get('street', '').strip() # Get street
        trade_type  = request.form.get('trade_type', '').strip() # Get trade type

        # Basic validation: check if required fields are empty
        if not full_name or not phone or not street_name or not trade_type:
            flash('Name, phone, street, and trade type are required fields.', 'danger')
            return redirect(url_for('register_vendor'))

        # Regex validation for phone number (ensures it's numeric and 10-15 digits)
        import re
        if not re.match(r'^\+?[0-9]{10,15}$', phone):
            flash('Invalid phone number format. Please use 10-15 digits.', 'danger')
            return redirect(url_for('register_vendor'))
            
        # Generate a unique Vendor ID (e.g. VND-001)
        count     = Vendor.query.count() + 1 # Total vendors + 1
        vendor_id = f'VND-{count:03d}' # Format as VND-###

        # Check if the street already exists in the streets table
        street_obj = Street.query.filter_by(name=street_name).first()
        if not street_obj: # If street is new, create a new record for it
            street_obj = Street(name=street_name)
            db.session.add(street_obj)
            db.session.flush() # Get the new street's ID before committing

        # Create the Vendor object with all provided data
        vendor = Vendor(
            vendor_id    = vendor_id,
            full_name    = full_name,
            phone        = phone,
            nin          = request.form.get('nin', ''), # National ID
            dob          = request.form.get('dob', ''), # Date of Birth
            street_id    = street_obj.id, # Link to street record
            street       = street_name, # Fallback name
            trade_type   = trade_type,
            stall_number = request.form.get('stall_number', ''),
            permit_start = request.form.get('permit_start', ''),
            permit_end   = request.form.get('permit_end', ''),
            status       = 'Active', # Default status
            notes        = request.form.get('notes', '')
        )
        db.session.add(vendor) # Save vendor
        db.session.commit() # Commit to database
        flash(f'Vendor {vendor.full_name} registered successfully! ID: {vendor_id}', 'success')
        return redirect(url_for('vendors'))

    return render_template('register_vendor.html', streets=streets)


@app.route('/vendors/<int:vendor_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'field_officer')
def edit_vendor(vendor_id):
    """Allows editing of an existing vendor's profile."""
    vendor = Vendor.query.get_or_404(vendor_id) # Find vendor or show error
    streets = Street.query.order_by(Street.name).all() # Get streets
    if request.method == 'POST':
        full_name   = request.form.get('full_name', '').strip()
        phone       = request.form.get('phone', '').strip()
        street_name = request.form.get('street', '').strip()
        trade_type  = request.form.get('trade_type', '').strip()
        
        # Validation for edit form
        if not full_name or not phone or not street_name or not trade_type:
            flash('Name, phone, street, and trade type are required fields.', 'danger')
            return redirect(url_for('edit_vendor', vendor_id=vendor.id))

        import re
        if not re.match(r'^\+?[0-9]{10,15}$', phone):
            flash('Invalid phone number format. Please use 10-15 digits.', 'danger')
            return redirect(url_for('edit_vendor', vendor_id=vendor.id))
            
        # Update the vendor object with new values
        vendor.full_name    = full_name
        vendor.phone        = phone
        vendor.nin          = request.form.get('nin', '')
        
        # Update street if it changed
        street_obj = Street.query.filter_by(name=street_name).first()
        if not street_obj:
            street_obj = Street(name=street_name)
            db.session.add(street_obj)
            db.session.flush()
        vendor.street_id = street_obj.id
        vendor.street    = street_name
        
        vendor.trade_type   = trade_type
        vendor.stall_number = request.form.get('stall_number', '')
        vendor.status       = request.form['status'] # Active/Pending/Expired
        vendor.notes        = request.form.get('notes', '')
        db.session.commit() # Save updates
        flash('Vendor updated successfully.', 'success')
        return redirect(url_for('vendor_detail', vendor_id=vendor.id))
    return render_template('edit_vendor.html', vendor=vendor, streets=streets)


# ── RECEIPT MANAGEMENT ──────────────────────────────────────────────────────
@app.route('/receipts')
@login_required
def receipts():
    """Lists all payment receipts with advanced filtering options."""
    search    = request.args.get('search', '') # Search keyword
    fee_type  = request.args.get('fee_type', '') # Fee type filter
    date_from = request.args.get('date_from', '') # Start date
    date_to   = request.args.get('date_to', '') # End date
    
    # Start query by joining Vendor table (so we can search by vendor name)
    query = Receipt.query.join(Vendor, Receipt.vendor_id == Vendor.id)
    
    # Data Isolation: Vendors can only see their OWN receipts
    if current_user.role == 'vendor' and current_user.vendor_ptr:
        query = query.filter(Receipt.vendor_id == current_user.vendor_ptr)
    
    # Apply keyword search
    if search:
        query = query.filter(
            db.or_(
                Receipt.receipt_no.ilike(f'%{search}%'),
                Vendor.full_name.ilike(f'%{search}%'),
                Vendor.vendor_id.ilike(f'%{search}%')
            )
        )
    
    # Apply fee type filter
    if fee_type:
        query = query.filter(Receipt.fee_type == fee_type)
        
    # Apply date filters
    if date_from:
        query = query.filter(Receipt.date_issued >= date_from)
    if date_to:
        query = query.filter(Receipt.date_issued <= date_to + " 23:59:59")
    
    # Finalize query: newest first
    receipt_list = query.order_by(Receipt.id.desc()).all()
    
    # Get unique fee types for the filter dropdown
    fee_types = db.session.query(Receipt.fee_type).distinct().all()
    fee_types = [f[0] for f in fee_types]
    
    return render_template('receipts.html', 
                           receipts=receipt_list, 
                           search=search, 
                           fee_type=fee_type,
                           date_from=date_from,
                           date_to=date_to,
                           fee_types=fee_types)


@app.route('/receipts/issue', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'field_officer') # Finance officers can't issue receipts
def issue_receipt():
    """Form to manually issue a receipt to an active vendor."""
    vendors = Vendor.query.filter_by(status='Active').all() # Only active vendors
    if request.method == 'POST':
        vendor_id_str = request.form.get('vendor_id', '').strip()
        fee_type      = request.form.get('fee_type', '').strip()
        amount_str    = request.form.get('amount', '').strip()
        
        # Validation
        if not vendor_id_str or not fee_type or not amount_str:
            flash('All required fields must be filled.', 'danger')
            return redirect(url_for('issue_receipt'))
            
        try:
            vendor_id = int(vendor_id_str)
            amount    = float(amount_str)
            if amount <= 0:
                raise ValueError("Amount must be positive.")
        except ValueError:
            flash('Invalid input: amount must be a positive number.', 'danger')
            return redirect(url_for('issue_receipt'))
            
        vendor = Vendor.query.get_or_404(vendor_id) # Confirm vendor exists

        # Generate a unique Receipt Number (RCP-####)
        count      = Receipt.query.count() + 1
        receipt_no = f'RCP-{count:04d}'

        # Create the receipt record
        receipt = Receipt(
            receipt_no   = receipt_no,
            vendor_id    = vendor.id,
            fee_type     = fee_type,
            amount       = amount,
            date_issued  = datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            issued_by    = current_user.username,
            notes        = request.form.get('notes', ''),
            status       = 'Paid' # Manually issued receipts are paid by default
        )
        db.session.add(receipt)
        db.session.commit() # Save
        flash(f'Receipt {receipt_no} issued for {vendor.full_name}.', 'success')
        return redirect(url_for('receipt_detail', receipt_id=receipt.id))

    return render_template('issue_receipt.html', vendors=vendors)


@app.route('/receipts/<int:receipt_id>/verify', methods=['POST'])
@login_required
@role_required('admin', 'finance_officer') # Finance oversight route
def verify_receipt_finance(receipt_id):
    """Allows Finance Officers to 'Verify' a transaction as audited."""
    receipt = Receipt.query.get_or_404(receipt_id)
    if not receipt.is_verified:
        receipt.is_verified = True # Mark as verified
        receipt.verified_by = current_user.username # Track who verified it
        receipt.verified_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db.session.commit()
        flash(f'Transaction {receipt.receipt_no} verified successfully.', 'success')
    else:
        flash('Transaction is already verified.', 'info')
    return redirect(request.referrer or url_for('receipts'))

@app.route('/receipts/<int:receipt_id>')
@login_required
def receipt_detail(receipt_id):
    """Shows the printable detail view of a receipt with a QR code."""
    receipt = Receipt.query.get_or_404(receipt_id)
    vendor  = Vendor.query.get(receipt.vendor_id)
    
    # ── QR CODE GENERATION ──────────────────────────────────────────────
    import qrcode
    import base64
    # Create the verification URL for the QR code
    verify_url = url_for('verify_receipt', receipt_no=receipt.receipt_no, _external=True)
    qr = qrcode.QRCode(version=1, box_size=4, border=1)
    qr.add_data(verify_url)
    qr.make(fit=True)
    buf = io.BytesIO() # Save to memory buffer
    qr.make_image(fill_color="black", back_color="white").save(buf, format='PNG')
    # Encode as Base64 to embed directly in the HTML <img> tag
    qr_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')
    
    return render_template('receipt_detail.html', receipt=receipt, vendor=vendor, qr_b64=qr_b64)


# ── FEE TRACKING SYSTEM ─────────────────────────────────────────────────────
@app.route('/fee-tracker')
@login_required
def fee_tracker():
    """Real-time matrix showing who has paid their daily fees for today."""
    today_str = date.today().strftime('%Y-%m-%d')
    schedules = FeeSchedule.query.all() # Get daily/weekly fee rates
    # Only active vendors appear on the tracker
    active_vendors = Vendor.query.filter(Vendor.status.in_(['Active', 'NotPaid'])).all()
    
    # Pre-fetch today's payments for fast lookup
    payments_list = FeePayment.query.filter_by(period_date=today_str).all()
    payments = {(p.vendor_id, p.schedule_id): p for p in payments_list}
    
    return render_template('fee_tracker.html', 
                           vendors=active_vendors, 
                           schedules=schedules, 
                           payments=payments,
                           today=today_str)


@app.route('/fee-tracker/pay', methods=['POST'])
@login_required
@role_required('admin', 'field_officer') # Collectors only
def mark_paid():
    """Marks a vendor's fee as paid for today and generates a receipt."""
    vendor_id   = int(request.form['vendor_id'])
    schedule_id = int(request.form['schedule_id'])
    today_str   = date.today().strftime('%Y-%m-%d')
    
    vendor   = Vendor.query.get_or_404(vendor_id)
    schedule = FeeSchedule.query.get_or_404(schedule_id)
    
    # Prevent double payment recording
    existing = FeePayment.query.filter_by(
        vendor_id=vendor_id, 
        schedule_id=schedule_id, 
        period_date=today_str,
        status='Paid'
    ).first()
    
    if existing:
        flash('Fee already recorded as paid.', 'info')
        return redirect(url_for('fee_tracker'))
    
    # STEP 1: Generate the Payment Receipt
    count      = Receipt.query.count() + 1
    receipt_no = f'RCP-{count:04d}'
    
    receipt = Receipt(
        receipt_no   = receipt_no,
        vendor_id    = vendor.id,
        fee_type     = schedule.name, # e.g. 'Daily Pitch Fee'
        amount       = schedule.amount,
        date_issued  = datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        issued_by    = current_user.username,
        status       = 'Paid'
    )
    db.session.add(receipt)
    db.session.flush() # Get the new receipt.id for the payment record
    
    # STEP 2: Record the FeePayment (the cell in the tracker)
    payment = FeePayment.query.filter_by(
        vendor_id=vendor_id, 
        schedule_id=schedule_id, 
        period_date=today_str
    ).first()
    
    if not payment: # Create new record if none exists for today
        payment = FeePayment(
            vendor_id=vendor_id,
            schedule_id=schedule_id,
            period_date=today_str
        )
        db.session.add(payment)
    
    payment.paid_date  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    payment.status     = 'Paid'
    payment.receipt_id = receipt.id # Link tracker record to physical receipt
    
    # Update vendor status: if they were 'NotPaid', mark them 'Active' again
    if schedule.frequency == 'daily':
        vendor.status = 'Active'
        
    db.session.commit() # Save all changes
    flash(f'Payment of UGX {schedule.amount:,.0f} recorded for {vendor.full_name}.', 'success')
    return redirect(url_for('fee_tracker'))


# ── FINES & VIOLATIONS ──────────────────────────────────────────────────────
@app.route('/fines')
@login_required
@role_required('admin', 'finance_officer', 'field_officer', 'vendor')
def fines():
    """Lists all penalty fines in the system."""
    query = Fine.query.order_by(Fine.id.desc())
    # Isolation: Vendors see only their own fines
    if current_user.role == 'vendor' and current_user.vendor_ptr:
        query = query.filter(Fine.vendor_id == current_user.vendor_ptr)
    
    fines_list = query.all()
    return render_template('fines.html', fines=fines_list)

@app.route('/fines/issue', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'field_officer') # Finance cannot issue fines
def issue_fine():
    """Allows a Field Officer to penalize a vendor for a violation."""
    vendors = Vendor.query.all() # Get all vendors (can fine anyone)
    if request.method == 'POST':
        vendor_id = int(request.form['vendor_id'])
        amount    = float(request.form['amount'])
        reason    = request.form['reason']
        # Create the fine record
        fine = Fine(
            vendor_id=vendor_id,
            amount=amount,
            reason=reason,
            date_issued=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            status='Unpaid', # Starts as unpaid
            issued_by=current_user.username
        )
        db.session.add(fine)
        db.session.commit()
        flash('Fine issued successfully.', 'success')
        return redirect(url_for('fines'))
    return render_template('issue_fine.html', vendors=vendors)

@app.route('/fines/<int:fine_id>/pay', methods=['POST'])
@login_required
@role_required('admin', 'finance_officer', 'field_officer')
def pay_fine(fine_id):
    """Processes payment for an outstanding fine and generates a receipt."""
    fine = Fine.query.get_or_404(fine_id)
    if fine.status == 'Paid':
        flash('Fine is already paid.', 'info')
        return redirect(url_for('fines'))
        
    fine.status = 'Paid' # Mark fine as resolved
    
    # Generate receipt for the fine payment
    count      = Receipt.query.count() + 1
    receipt_no = f'RCP-{count:04d}'
    
    receipt = Receipt(
        receipt_no   = receipt_no,
        vendor_id    = fine.vendor_id,
        fee_type     = f'Fine: {fine.reason}', # Explain what they paid for
        amount       = fine.amount,
        date_issued  = datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        issued_by    = current_user.username,
        status       = 'Paid'
    )
    db.session.add(receipt)
    db.session.commit() # Save both fine status and new receipt
    flash(f'Fine paid successfully! Receipt {receipt_no} generated.', 'success')
    return redirect(url_for('receipt_detail', receipt_id=receipt.id))


# ── DATA REPORTING ──────────────────────────────────────────────────────────
@app.route('/reports')
@login_required
@role_required('admin', 'finance_officer') # Only financial roles can see reports
def reports():
    """Generates visual analytics and revenue reports."""
    from sqlalchemy import func
    
    # 1. Calculate Daily revenue for the last 7 days
    today = date.today()
    last_7_days = []
    daily_revenue = []
    for i in range(6, -1, -1): # Loop backwards from 6 days ago to today
        d = today - timedelta(days=i)
        d_str = d.strftime('%Y-%m-%d')
        # Sum all receipt amounts matching this date
        rev = db.session.query(func.sum(Receipt.amount)).filter(Receipt.date_issued.like(d_str + '%')).scalar() or 0
        last_7_days.append(d.strftime('%a %d')) # Label (e.g. 'Mon 25')
        daily_revenue.append(rev) # Data value
        
    # 2. Get Top 5 vendors by total contribution
    top_vendors = db.session.query(Vendor.full_name, func.sum(Receipt.amount))\
                    .join(Receipt).group_by(Vendor.id)\
                    .order_by(func.sum(Receipt.amount).desc()).limit(5).all()
                    
    # 3. Get Revenue distribution by street/zone
    street_revenue = db.session.query(Street.name, func.sum(Receipt.amount))\
                        .join(Vendor, Vendor.street_id == Street.id)\
                        .join(Receipt, Receipt.vendor_id == Vendor.id)\
                        .group_by(Street.id).order_by(func.sum(Receipt.amount).desc()).all()

    # Summary numeric stats
    today_str = today.strftime('%Y-%m-%d')
    stats = {
        'today_total': db.session.query(func.sum(Receipt.amount)).filter(Receipt.date_issued.like(today_str + '%')).scalar() or 0,
        'week_total':  db.session.query(func.sum(Receipt.amount)).filter(Receipt.date_issued >= (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')).scalar() or 0,
        'paid_today_count': FeePayment.query.filter_by(period_date=today_str, status='Paid').count()
    }

    return render_template('reports.html', 
                           daily_labels=last_7_days, 
                           daily_values=daily_revenue,
                           top_vendors=top_vendors,
                           street_revenue=street_revenue,
                           stats=stats)

@app.route('/api/reports')
@login_required
@role_required('admin', 'finance_officer')
def api_reports():
    """Returns report data in JSON format for the interactive charts."""
    from sqlalchemy import func
    today = date.today()
    
    # Build revenue trend array
    daily_revenue = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        d_str = d.strftime('%Y-%m-%d')
        rev = db.session.query(func.sum(Receipt.amount)).filter(Receipt.date_issued.like(d_str + '%')).scalar() or 0
        daily_revenue.append({'date': d_str, 'value': rev})
        
    # Get street distribution
    street_rev = db.session.query(Street.name, func.sum(Receipt.amount))\
                    .join(Vendor, Vendor.street_id == Street.id)\
                    .join(Receipt, Receipt.vendor_id == Vendor.id)\
                    .group_by(Street.id).all()
    
    return jsonify({
        'revenue_trend': daily_revenue,
        'street_distribution': [{'name': s, 'total': t} for s, t in street_rev]
    })



@app.route('/api/vendors')
@login_required
def api_vendors():
    """Return all vendors as JSON."""
    vendor_list = Vendor.query.all()
    return jsonify([v.to_dict() for v in vendor_list])


@app.route('/api/stats')
@login_required
def api_stats():
    """Return system stats as JSON."""
    return jsonify({
        'total_vendors':  Vendor.query.count(),
        'active_vendors': Vendor.query.filter_by(status='Active').count(),
        'total_receipts': Receipt.query.count(),
        'total_revenue':  db.session.query(db.func.sum(Receipt.amount)).scalar() or 0
    })


# ── PDF GENERATION (PHASE 2) ────────────────────────────────────────────────
@app.route('/receipts/<int:receipt_id>/pdf')
@login_required
def receipt_pdf(receipt_id):
    """Generates an official PDF receipt for download."""
    if not REPORTLAB_OK: # Check if library is installed
        flash('PDF generation is not available. Install reportlab.', 'danger')
        return redirect(url_for('receipt_detail', receipt_id=receipt_id))

    receipt = Receipt.query.get_or_404(receipt_id)
    vendor  = Vendor.query.get(receipt.vendor_id)

    buf    = io.BytesIO() # Buffer to store PDF data in memory
    doc    = SimpleDocTemplate(buf, pagesize=A5,
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = [] # List of elements (text, tables, etc.) in the PDF

    # --- Define convenience styles ---
    center = ParagraphStyle('center', parent=styles['Normal'],
                            alignment=1, spaceAfter=2)
    bold   = ParagraphStyle('bold', parent=styles['Normal'],
                            fontName='Helvetica-Bold', fontSize=9)
    small  = ParagraphStyle('small', parent=styles['Normal'],
                            fontSize=8, textColor=colors.grey)

    # Header branding
    import os
    from reportlab.platypus import Image
    logo_path = os.path.join(app.root_path, 'static', 'img', 'logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path, width=3*cm, height=3*cm)
        img.hAlign = 'CENTER'
        img.spaceAfter = 10
        story.append(img)
        
    story.append(Paragraph('<b>Arua City Council</b>', ParagraphStyle(
        'h1', parent=styles['Normal'], alignment=1, fontSize=13,
        fontName='Helvetica-Bold', spaceAfter=2)))
    story.append(Paragraph('Street Vendor Authority', center))
    story.append(Paragraph('Arua, Uganda', small))
    story.append(HRFlowable(width='100%', thickness=1, spaceAfter=8))

    # Data table construction
    def row(label, value):
        return [Paragraph(f'<b>{label}</b>', styles['Normal']),
                Paragraph(str(value), styles['Normal'])]

    data = [
        row('Receipt No.', receipt.receipt_no),
        row('Date',        receipt.date_issued[:10]),
        row('Time',        receipt.date_issued[11:16] if len(receipt.date_issued) > 10 else '—'),
        row('Vendor Name', vendor.full_name if vendor else '—'),
        row('Vendor ID',   vendor.vendor_id if vendor else '—'),
        row('Street/Zone', vendor.street    if vendor else '—'),
        row('Fee Type',    receipt.fee_type),
        row('Issued By',   receipt.issued_by),
    ]
    if receipt.notes:
        data.append(row('Notes', receipt.notes))

    tbl = Table(data, colWidths=[4.5*cm, 8*cm])
    tbl.setStyle(TableStyle([
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.whitesmoke, colors.white]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.lightgrey),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width='100%', thickness=1, lineCap='round', spaceAfter=6))

    # Total amount summary
    total_style = ParagraphStyle('total', parent=styles['Normal'],
                                 fontSize=12, fontName='Helvetica-Bold', alignment=2)
    story.append(Paragraph(f'TOTAL PAID: UGX {receipt.amount:,.0f}', total_style))
    story.append(Spacer(1, 8))

    status_color = '#1D9E75' if receipt.status == 'Paid' else '#D85A30'
    story.append(Paragraph(
        f'<font color="{status_color}"><b>Status: {receipt.status}</b></font>', center))
        
    # Generate and embed QR code for verification
    import qrcode
    verify_url = url_for('verify_receipt', receipt_no=receipt.receipt_no, _external=True)
    qr = qrcode.QRCode(version=1, box_size=5, border=1)
    qr.add_data(verify_url)
    qr.make(fit=True)
    qr_buf = io.BytesIO()
    qr.make_image(fill_color="black", back_color="white").save(qr_buf, format='PNG')
    qr_buf.seek(0)
    
    from reportlab.platypus import Image as RLImage
    qr_img = RLImage(qr_buf, width=2.5*cm, height=2.5*cm)
    qr_img.hAlign = 'CENTER'
    qr_img.spaceBefore = 8
    story.append(qr_img)

    story.append(Paragraph(f'<font color="grey"><i>{receipt.receipt_no} — scan to verify</i></font>',
                           ParagraphStyle('barcode', parent=styles['Normal'],
                                         alignment=1, fontSize=8, textColor=colors.grey)))

    doc.build(story) # Construct the file
    buf.seek(0)
    # Return as PDF response
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{receipt.receipt_no}.pdf"'
    return response



@app.route('/vendors/<int:vendor_id>/permit/pdf')
@login_required
def vendor_permit_pdf(vendor_id):
    """Generates an official printable Vendor Permit PDF."""
    if not REPORTLAB_OK:
        flash('PDF generation is not available. Install reportlab.', 'danger')
        return redirect(url_for('vendor_detail', vendor_id=vendor_id))

    vendor = Vendor.query.get_or_404(vendor_id) # Find vendor

    buf    = io.BytesIO() # Memory buffer
    from reportlab.lib.pagesizes import landscape, A6
    doc    = SimpleDocTemplate(buf, pagesize=landscape(A6),
                               leftMargin=1*cm, rightMargin=1*cm,
                               topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story  = []
    
    # Logo header
    import os
    from reportlab.platypus import Image as RLImage
    logo_path = os.path.join(app.root_path, 'static', 'img', 'logo.png')
    if os.path.exists(logo_path):
        img_logo = RLImage(logo_path, width=1.5*cm, height=1.5*cm)
        img_logo.hAlign = 'CENTER'
        img_logo.spaceAfter = 5
        story.append(img_logo)
        
    story.append(Paragraph('<b>OFFICIAL VENDOR PERMIT</b>', ParagraphStyle(
        'h1', parent=styles['Normal'], alignment=1, fontSize=12,
        fontName='Helvetica-Bold', spaceAfter=10, textColor='#2B5EAD')))

    # Data row helper
    def row(label, value):
        return [Paragraph(f'<b>{label}</b>', styles['Normal']),
                Paragraph(str(value), styles['Normal'])]

    # Permit data
    data = [
        row('Vendor Name:', vendor.full_name),
        row('Vendor ID:', vendor.vendor_id),
        row('Location:', vendor.street),
        row('Trade Type:', vendor.trade_type),
        row('Valid Until:', vendor.permit_end),
    ]

    tbl = Table(data, colWidths=[3*cm, 7*cm])
    tbl.setStyle(TableStyle([
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    
    # Generate verification QR
    import qrcode
    verify_url = url_for('verify_vendor', vendor_id=vendor.id, _external=True)
    qr = qrcode.QRCode(version=1, box_size=5, border=1)
    qr.add_data(verify_url)
    qr.make(fit=True)
    qr_buf = io.BytesIO()
    qr.make_image(fill_color="black", back_color="white").save(qr_buf, format='PNG')
    qr_buf.seek(0)
    
    qr_img = RLImage(qr_buf, width=2.5*cm, height=2.5*cm)
    
    # Layout permit table and QR side-by-side
    layout_data = [[tbl, qr_img]]
    layout_tbl = Table(layout_data, colWidths=[10*cm, 3*cm])
    layout_tbl.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    story.append(layout_tbl)

    story.append(Paragraph(f'<font color="grey"><i>Scan to verify authenticity</i></font>',
                           ParagraphStyle('barcode_text', parent=styles['Normal'],
                                         alignment=2, fontSize=7, textColor=colors.grey)))

    doc.build(story)
    buf.seek(0)
    # Send PDF file to browser
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="permit_{vendor.vendor_id}.pdf"'
    return response


# ── EXCEL EXPORTS (PHASE 2) ───────────────────────────────────────────────────
@app.route('/vendors/export')
@login_required
def export_vendors():
    """Exports the entire vendor registry to an Excel spreadsheet."""
    if not OPENPYXL_OK:
        flash('Excel export is not available. Install openpyxl.', 'danger')
        return redirect(url_for('vendors'))

    wb = openpyxl.Workbook() # Create new workbook
    ws = wb.active # Get active sheet
    ws.title = 'Vendors'

    # Column headers
    headers = ['Vendor ID', 'Full Name', 'Phone', 'National ID', 'Street/Zone',
               'Trade Type', 'Stall No.', 'Status', 'Permit Start', 'Permit End']
    ws.append(headers)

    # Style the header row (Blue background, White bold text)
    header_fill = PatternFill('solid', fgColor='2B5EAD')
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add all vendors from database
    for v in Vendor.query.order_by(Vendor.id).all():
        ws.append([v.vendor_id, v.full_name, v.phone, v.nin,
                   v.street, v.trade_type, v.stall_number,
                   v.status, v.permit_start, v.permit_end])

    # Auto-adjust column widths for readability
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf) # Save to memory
    buf.seek(0)
    response = make_response(buf.read())
    # Set headers for Excel file download
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename="vendors.xlsx"'
    return response


@app.route('/receipts/export')
@login_required
def export_receipts():
    """Exports the transaction history to an Excel spreadsheet."""
    if not OPENPYXL_OK:
        flash('Excel export is not available. Install openpyxl.', 'danger')
        return redirect(url_for('receipts'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Receipts'

    headers = ['Receipt #', 'Vendor Name', 'Vendor ID', 'Fee Type',
               'Amount (UGX)', 'Date Issued', 'Issued By', 'Status']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='2B5EAD')
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add all receipts with linked vendor info
    for r in Receipt.query.order_by(Receipt.id).all():
        vendor_name = r.vendor_ref.full_name if r.vendor_ref else ''
        vendor_id   = r.vendor_ref.vendor_id if r.vendor_ref else ''
        ws.append([r.receipt_no, vendor_name, vendor_id, r.fee_type,
                   r.amount, r.date_issued, r.issued_by, r.status])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename="receipts.xlsx"'
    return response



# ── Phase 2: SMS Reminder ────────────────────────────────────────────────────
import warnings
from urllib3.exceptions import InsecureRequestWarning
warnings.simplefilter('ignore', InsecureRequestWarning)

import requests
if not hasattr(requests.Session, '_ssl_override'):
    _orig_req = requests.Session.request
    def _patched_req(self, method, url, **kwargs):
        kwargs['verify'] = False
        return _orig_req(self, method, url, **kwargs)
    requests.Session.request = _patched_req
    requests.Session._ssl_override = True

def send_sms(phone, message):
    """Send SMS via Africa's Talking (sandbox by default). Returns (ok, info_str)."""
    if not AFRICASTALKING_OK:
        return False, 'africastalking package not installed'
    try:
        africastalking.initialize(AT_USERNAME, AT_API_KEY)
        sms = africastalking.SMS
        recipients = [phone] if not isinstance(phone, list) else phone
        kwargs = dict(message=message, recipients=recipients)
        if AT_SENDER:
            kwargs['sender_id'] = AT_SENDER
        result = sms.send(**kwargs)
        return True, str(result)
    except Exception as exc:
        return False, str(exc)


@app.route('/vendors/<int:vendor_id>/sms', methods=['POST'])
@login_required
def send_sms_reminder(vendor_id):
    vendor  = Vendor.query.get_or_404(vendor_id)
    if not vendor.phone:
        flash('Vendor has no phone number on record.', 'warning')
        return redirect(url_for('vendor_detail', vendor_id=vendor_id))

    permit_info = f'Permit expires: {vendor.permit_end}.' if vendor.permit_end else ''
    message = (
        f'Dear {vendor.full_name}, this is a reminder from Arua City '
        f'Council – Street Vendor Authority. Please ensure your vendor '
        f'registration fees are up to date. {permit_info} '
        f'Contact us on +256 417 123456 for assistance.'
    )
    ok, info = send_sms(vendor.phone, message)
    if ok:
        flash(f'SMS reminder sent to {vendor.phone}.', 'success')
    else:
        flash(f'SMS could not be sent: {info}', 'warning')
    return redirect(url_for('vendor_detail', vendor_id=vendor_id))

# ── Verification System ──────────────────────────────────────────────────────
@app.route('/verify/receipt/<receipt_no>')
def verify_receipt(receipt_no):
    receipt = Receipt.query.filter_by(receipt_no=receipt_no).first()
    if not receipt:
        return render_template('verify_result.html', receipt=None, valid=False, entity="Receipt")
    vendor = Vendor.query.get(receipt.vendor_id)
    return render_template('verify_result.html', receipt=receipt, vendor=vendor, valid=True, entity="Receipt")

@app.route('/verify/vendor/<int:vendor_id>')
def verify_vendor(vendor_id):
    vendor = Vendor.query.get(vendor_id)
    if not vendor:
        return render_template('verify_result.html', vendor=None, valid=False, entity="Vendor Permit")
    return render_template('verify_result.html', vendor=vendor, valid=True, entity="Vendor Permit")




# ── SYSTEM INITIALIZATION ───────────────────────────────────────────────────
def seed_demo_data():
    """Initializes the database with essential system accounts and configurations."""
    # 1. Create Default Administrative Users if none exist
    if User.query.count() == 0:
        admin = User(username='admin', email='admin@arua.go.ug', role='admin')
        admin.set_password('admin123')
        
        field = User(username='field', email='field@arua.go.ug', role='field_officer')
        field.set_password('field123')
        
        finance = User(username='finance', email='finance@arua.go.ug', role='finance_officer')
        finance.set_password('finance123')
        
        db.session.add_all([admin, field, finance])

    # 2. Pre-populate Primary Streets/Zones
    if Street.query.count() == 0:
        sample_streets = [
            Street(name='Arua Avenue'),
            Street(name='Hospital Rd.'),
            Street(name='Market Lane'),
            Street(name='Onduparaka Rd.'),
            Street(name='Blue Tower St.')
        ]
        db.session.add_all(sample_streets)
        db.session.flush()

    # 3. Define Standard Fee Schedules
    if FeeSchedule.query.count() == 0:
        schedules = [
            FeeSchedule(name='Daily Pitch Fee', amount=12000, frequency='daily'),
            FeeSchedule(name='Weekly Permit',    amount=50000, frequency='weekly')
        ]
        db.session.add_all(schedules)
        db.session.flush()

    # Commit all structural initialization
    db.session.commit()



# ── SYSTEM ENTRY POINT ────────────────────────────────────────────────────────
if __name__ == '__main__':
    """Main application entry point."""
    with app.app_context():
        db.create_all() # Create database tables if they do not exist
        seed_demo_data() # Add default/sample data to the database
    # Run the Flask development server on the local machine
    app.run(debug=True)

